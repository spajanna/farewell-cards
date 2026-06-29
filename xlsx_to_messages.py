import json
import sys
from openpyxl import load_workbook

# -----------------------------------------------
# Anpassen: Spaltennamen aus dem MS Forms Export
# -----------------------------------------------
SPALTE_NAME      = "Vollständiger Name (der für Maria angezeigt werden soll)"
SPALTE_NACHRICHT = "Deine persönliche Nachricht\xa0"
# -----------------------------------------------
# Datei einlesen
# -----------------------------------------------
xlsx_datei = "antworten.xlsx"

if len(sys.argv) > 1:
    xlsx_datei = sys.argv[1]

wb = load_workbook(xlsx_datei)
ws = wb.active

# Kopfzeile einlesen
headers = [cell.value for cell in ws[1]]
print("Gefundene Spalten:", headers)

# Spaltenindizes ermitteln
try:
    idx_name      = headers.index(SPALTE_NAME)
    idx_nachricht = headers.index(SPALTE_NACHRICHT)
except ValueError as e:
    print(f"\nFehler: Spalte nicht gefunden – {e}")
    print("Bitte SPALTE_NAME / SPALTE_NACHRICHT im Script anpassen.")
    sys.exit(1)

# -----------------------------------------------
# Zeilen einlesen
# -----------------------------------------------
messages = []

for row in ws.iter_rows(min_row=2, values_only=True):
    name      = row[idx_name]
    nachricht = row[idx_nachricht]

    if name and nachricht:
        messages.append({
            "name":    str(name).strip(),
            "message": str(nachricht).strip()
        })

# -----------------------------------------------
# Ausgabe
# -----------------------------------------------
js_output = "const messages = " + json.dumps(messages, ensure_ascii=False, indent=2) + ";"

print("\n--- Kopiere das hier in deine HTML-Datei ---\n")
print(js_output)

with open("messages_output.js", "w", encoding="utf-8") as f:
    f.write(js_output)

print("\n--- Auch gespeichert als messages_output.js ---")
